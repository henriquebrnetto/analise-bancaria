from ast import pattern
from calendar import month_abbr
from ctypes import sizeof
from genericpath import isfile
from msilib.schema import Font
from venv import create
from xml.etree.ElementInclude import include
import pandas as pd
from project_func import create_book, file_reader
import numpy as np
from datetime import datetime, date
import os, re, sys, shutil
from openpyxl.styles import Font

def main():
    #Get files from Download folder
    src_path = 'C:\\Downloads'
    path = 'D:\\Python Projects\\bank analysis\\extratos'
    download_files = []
    for file in os.listdir(src_path):
        if file.endswith('.csv'):
            download_files.append(f'{src_path}\\{file}')
    
    #Move files to the destination folder
    for file_name in download_files:
        shutil.move(file_name, path)

    #Get all Files on Folder
    filenames = []
    for file in os.listdir(path):
        filenames.append(f'{path}\\{file}')
        
    #Read files
    df_list = file_reader(filenames)

    try:
        df = pd.concat(df_list, ignore_index=True)
    except ValueError as e:
        print(f'{type(e).__name__} : There are no files in the destination folder')
        sys.exit(1)

    #Transform Date Strings to Date type
    df['Data'] = pd.to_datetime(df['Data'], dayfirst=True, infer_datetime_format=True)
    df['Data'] = df['Data'].apply(lambda x: x.date())
    
    #Drop extra column
    df.drop(axis=1, columns='Unnamed: 6', inplace=True)

    #Create New Columns for Classification
    df['Classificação'] = np.nan
    df['Descrição'] = np.nan

    #Select Start of the Month Vals
    month_start_vals = df.loc[df['Histórico'] == 'Saldo Anterior']
    month_start_vals = month_start_vals[['Data', 'Valor']].values

    #Select End of the Month Indexes and Vals
    month_end_vals = df.loc[df['Histórico'] == 'S A L D O']
    month_end_vals = month_end_vals[['Data', 'Valor']].values

    #DataFrame for Month Summary
    gastos_df = pd.DataFrame(columns = ['Mês','Salário', 'Total Gasto', 'Saldo Final', '+/-'])

    #Append Results
    for index, val in enumerate(month_start_vals):
        month_e = date.strftime(month_end_vals[index,0], '%B')
        gastos_df.loc[len(gastos_df.index)+1, :] = [month_e, None, None, month_end_vals[index][1], None]

    #Free Memory
    del month_end_vals, month_start_vals, month_e, df_list

    #Remove Start and End of the Month Vals from Main DF
    df = (df[(df['Histórico'] != 'Saldo Anterior') & (df['Histórico'] != 'S A L D O')]).reset_index(drop=True)

    #Split into Money Spent and Received
    val_in = df[df['Valor']>=0]
    val_out = df[df['Valor']<0]

    #Find Total Spent, Total Received and "+/-" of the Month
    for val in gastos_df['Mês'].items():
        month_num = datetime.strptime(val[1], '%B').month
        pattern = re.compile(f'2022-0{month_num}-\d')
        mon_sum_list = []
        exp_list = []
        for i, v in df['Data'].items():
            p = pattern.search(str(v))
            if p != None:
                if i in val_in.index:
                    mon_sum_list.append(val_in.loc[i,['Valor']][0])
                if i in val_out.index:
                    exp_list.append(val_out.loc[i,['Valor']][0])
        mon_sum = sum(mon_sum_list)
        tot_exp = sum(exp_list)
        total = mon_sum - abs(tot_exp)
        gastos_df.loc[val[0],['Total Gasto']] = abs(tot_exp)
        gastos_df.loc[val[0],['Salário']] = mon_sum
        gastos_df.loc[val[0],['+/-']] = total

    #DataFrames that will be transformed into different Excel Sheets
    pix = pd.DataFrame(columns= ['Data', 'Valor', 'Descrição'])
    seguros = pd.DataFrame(columns= ['Data', 'Valor', 'Descrição'])
    faturas = pd.DataFrame(columns= ['Data', 'Valor', 'Descrição'])
    cartao_credito = pd.DataFrame(columns= ['Mês','Data', 'Valor', 'Descrição'])
    investimento = pd.DataFrame(columns= ['Data', 'Valor', 'Descrição'])
    boletos = pd.DataFrame(columns= ['Data', 'Valor', 'Descrição'])
    ted = pd.DataFrame(columns= ['Data', 'Valor', 'Descrição'])

    #Create Dictionary with all Dataframes and Excel Sheet names (NOT USED, but may be useful in future updates on the script)
    final_df_feats = ['Data','Histórico','Valor','Classificação','Descrição']
    sheet_names = ['Extratos','Resumo Mensal','PIX','Seguros','Faturas','Cartao Credito','Investimentos','Boletos','TED']
    df_list = [df[final_df_feats],gastos_df,pix,seguros,faturas,cartao_credito,investimento,boletos,ted]
    dfs = dict(zip(sheet_names, df_list))
    
    #Description
    for i, row in val_out['Histórico'].items():
        #Pix e Subdivisões
        if 'PIX' in row.upper():
            #Pattern Selection
            p = re.search(r'\d{2}(:\d{2})', row)
            descr = row[p.span()[1]+1:]
            #Append values to DFs
            df.loc[i,['Descrição']] = f'Pix - {descr}'
            pix.loc[len(pix.index)] = [df.loc[i,['Data']].values[0], df.loc[i,['Valor']].values[0], df.loc[i,['Descrição']][0]]
        #Seguros
        if 'PLANO DE SAÚDE' in row.upper():
            df.loc[i,['Descrição']] = 'PLANO DE SAÚDE'
            seguros.loc[len(seguros.index)] = [df.loc[i,['Data']].values[0], df.loc[i,['Valor']].values[0], df.loc[i,['Descrição']][0]]
        #Contas
        if 'FATURA DE GÁS' in row.upper():
            df.loc[i,['Descrição']] = 'FATURA DE GÁS'
            faturas.loc[len(faturas.index)] = [df.loc[i,['Data']].values[0], df.loc[i,['Valor']].values[0], df.loc[i,['Descrição']][0]]
        if 'ENERGIA ELÉTRICA' in row.upper() or 'CONTA LUZ' in row.upper():
            if 'CPFL' in row.upper():
                df.loc[i,['Descrição']] = 'ENERGIA ELÉTRICA SN'
                faturas.loc[len(faturas.index)] = [df.loc[i,['Data']].values[0], df.loc[i,['Valor']].values[0], df.loc[i,['Descrição']][0]]
            elif abs(val_out.loc[i,['Valor']].values[0]) >= 80.00:
                df.loc[i,['Descrição']] = 'ENERGIA ELÉTRICA I'
                faturas.loc[len(faturas.index)] = [df.loc[i,['Data']].values[0], df.loc[i,['Valor']].values[0], df.loc[i,['Descrição']][0]]
            else:
                df.loc[i,['Descrição']] = 'ENERGIA ELÉTRICA H'
                faturas.loc[len(faturas.index)] = [df.loc[i,['Data']].values[0], df.loc[i,['Valor']].values[0], df.loc[i,['Descrição']][0]]
        if 'TELEFONE' in row.upper():
            if 'VIVO' in row.upper():
                df.loc[i,['Descrição']] = 'TELEFONE V'
                faturas.loc[len(faturas.index)] = [df.loc[i,['Data']].values[0], df.loc[i,['Valor']].values[0], df.loc[i,['Descrição']][0]]
            elif 'CLARO' in row.upper():
                df.loc[i,['Descrição']] = 'TELEFONE C'
                faturas.loc[len(faturas.index)] = [df.loc[i,['Data']].values[0], df.loc[i,['Valor']].values[0], df.loc[i,['Descrição']][0]]
            else:
                df.loc[i,['Descrição']] = 'TELEFONE'
                faturas.loc[len(faturas.index)] = [df.loc[i,['Data']].values[0], df.loc[i,['Valor']].values[0], df.loc[i,['Descrição']][0]]
        if 'CARTÃO CRÉDITO' in row.upper():
            month = df.loc[i,['Data']].values[0].strftime('%B')
            df.loc[i,['Descrição']] = 'CARTÃO DE CRÉDITO'
            cartao_credito.loc[len(cartao_credito.index)] = [month, df.loc[i,['Data']].values[0], df.loc[i,['Valor']].values[0], df.loc[i,['Descrição']][0]]
        if 'BB RF LP HIGH' in row.upper():
            df.loc[i,['Descrição']] = 'APLICAÇÃO'
            investimento.loc[len(investimento.index)] = [df.loc[i,['Data']].values[0], df.loc[i,['Valor']].values[0], df.loc[i,['Descrição']][0]]
        if 'BOLETO' in row.upper():
            #Pattern Selection
            p = re.search(r'- ', row)
            descr = row[p.span()[1]:]
            df.loc[i,['Descrição']] = f'BOLETO - {descr}'
            boletos.loc[len(boletos.index)] = [df.loc[i,['Data']].values[0], df.loc[i,['Valor']].values[0], df.loc[i,['Descrição']][0]]
        if 'TED' in row.upper():
            #Pattern Selection
            p = re.search(r'\b\d{10,15}\s', row)
            descr = row[p.span()[1]:]
            df.loc[i,['Descrição']] = f'TED - {descr}'
            ted.loc[len(boletos.index)] = [df.loc[i,['Data']].values[0], df.loc[i,['Valor']].values[0], df.loc[i,['Descrição']][0]]
        if 'TRANSFERÊNCIA PERIÓDICA' in row.upper():
            #Pattern Selection (before and after)
            p_in = re.search(r'\d{4,5}-\d{1}', row)
            p_fin = re.search(r'\d{3}/\d{3}', row)
            descr = row[p_in.span()[1]+1:p_fin.span()[0]]
            df.loc[i,['Descrição']] = descr
        if 'M J B R' in row.upper() or '8***-4 M' in row.upper():
            df.loc[i,['Descrição']] = 'TRANSF. M'
        if 'IMPOSTOS' in row.upper():
            df.loc[i,['Descrição']] = 'IMPOSTO'
        if 'COMPRA COM CARTÃO' in row.upper():
            p = re.search(r'\d{2}(:\d{2})', row)
            descr = row[p.span()[1]+1:]
            df.loc[i,['Descrição']] = descr
        if 'NET' in row.upper():
            df.loc[i,['Descrição']] = 'NET'
            faturas.loc[len(faturas.index)] = [df.loc[i,['Data']].values[0], df.loc[i,['Valor']].values[0], df.loc[i,['Descrição']][0]]
    
    #Sort DataFrame by type of Expense
    faturas = faturas.sort_values(by=['Descrição','Data'], ignore_index=True)

    #Writing Data to Excel
    #If file "Análise.xlsx" already exists
    if os.path.isfile('D:\\Python Projects\\bank analysis\\Análise.xlsx'):
        with pd.ExcelWriter('Análise.xlsx', engine='openpyxl', date_format='YYYY-MM-DD', mode='a', if_sheet_exists='overlay') as writer:
            df[final_df_feats].to_excel(writer, 'Extratos', header=False, index=False, startrow = writer.sheets['Extratos'].max_row)
            gastos_df.to_excel(writer, 'Resumo Mensal', header=False, index=False, startrow = writer.sheets['Resumo Mensal'].max_row)
            pix.to_excel(writer, 'Pix', header=False, index=False, startrow = writer.sheets['Pix'].max_row)
            faturas.to_excel(writer, 'Faturas', header=False, index=False, startrow = writer.sheets['Faturas'].max_row)
            seguros.to_excel(writer, 'Seguros', header=False, index=False, startrow = writer.sheets['Seguros'].max_row)
            cartao_credito.to_excel(writer, 'Cartão de Crédito', header=False, index=False, startrow = writer.sheets['Cartão de Crédito'].max_row)
            investimento.to_excel(writer, 'Aplicações', header=False, index=False, startrow = writer.sheets['Aplicações'].max_row)
            boletos.to_excel(writer, 'Boletos', header=False, index=False, startrow = writer.sheets['Boletos'].max_row)
            ted.to_excel(writer, 'TED', header=False, index=False, startrow = writer.sheets['TED'].max_row)

    #If file "Análise.xlsx" does not exist
    else:
        with pd.ExcelWriter('Análise.xlsx', engine='openpyxl', date_format='YYYY-MM-DD') as writer:
            df[final_df_feats].to_excel(writer, 'Extratos', index = False)
            gastos_df.to_excel(writer, 'Resumo Mensal', index = False)
            pix.to_excel(writer, 'Pix', index = False)
            faturas.to_excel(writer, 'Faturas', index = False)
            seguros.to_excel(writer, 'Seguros', index = False)
            cartao_credito.to_excel(writer, 'Cartão de Crédito', index = False)
            investimento.to_excel(writer, 'Aplicações', index = False)
            boletos.to_excel(writer, 'Boletos', index = False)
            ted.to_excel(writer, 'TED', index = False)
    
    #Colouring Negative values
    wb, ws = create_book(filename='Análise')
    sheet = wb['Resumo Mensal']
    for i in range(2, sheet.max_row+1):
        if int(sheet[f'E{i}'].value) < 0:
            sheet[f'E{i}'].font = Font(color='00FF0000')
    wb.save('Análise.xlsx')
    
    #DataFrame to see which cells do not have a description
    #df_missing = df[(pd.isna(df['Descrição']) & (df['Valor'] < 0))]

    #Create definite DF
    if os.path.isfile('D:\\Python Projects\\bank analysis\\extratos-df.csv'):
        df.to_csv('extratos-df.csv', mode = 'a', index=False, header = False, date_format='YYYY-MM-DD')
    else:
        df.to_csv('extratos-df.csv', index=False, date_format='YYYY-MM-DD')
    
    #Delete Used Files
    for file in filenames:
        os.remove(file)
    

if __name__ == '__main__':
    main()
